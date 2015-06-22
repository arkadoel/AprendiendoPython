from acciones import Acciones

__author__ = 'arkadoel'
import os
import pandas
from pandas import ExcelWriter
import xlsxwriter

#xlwt
from time import *
import xlwt

#openpyxl
import openpyxl


COL_NETO = 3
ROW_INIT = 5

class EscribirExcel():

    def __init__(self, salida="./salida.xlsx", entrada=""):
        """
        Iniciamos diciendo cual sera la ruta de salida
        """
        self.ruta_salida = salida
        self.EXCEL_DATOS = entrada

    def crear_un_libro_pandas_Excel_Writer(self):

        df = Acciones(ruta = self.EXCEL_DATOS).obtener_dataframe()
        df20 = df.head(20)
        #no existe archivo, lo creamos
        writer = pandas.ExcelWriter(self.ruta_salida)
        df20.to_excel(writer, sheet_name="20 lineas")

        writer.save()

    def crear_libro_xlwt(self):

        wb = xlwt.Workbook()
        hoja = wb.add_sheet('20 lineas')
        
        for i in range(0, 20):
            assert isinstance(hoja, xlwt.Worksheet )
            hoja.write(ROW_INIT + i, COL_NETO, "numero " + str(i))

        wb.save('salidaXLS.xls')

    def crear_libro_openpyxl(self):
        wb = openpyxl.Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = 'lol'

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        wb.save("salida2.xlsx")

    def modificar_con_openpyxl(self):
        from openpyxl.worksheet import Worksheet
        from openpyxl.workbook import Workbook
        from openpyxl.cell import Cell

        wb = openpyxl.load_workbook(filename="salida2.xlsx")
        ws = wb.get_sheet_by_name('lol')
        assert isinstance(ws, Worksheet)
        ws['B12'] = '21'

        #cambiamos el formato a texto
        ws['A2'] = 'l: ' + str(ws['A2'].value)
        c = ws['A2']
        assert  isinstance(c, Cell)
        c.font =

        ws['A3'] = 1234

        wb.save("salida2.xlsx")

